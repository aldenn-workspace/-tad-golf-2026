import openpyxl

path = "/Users/mini/Downloads/2025_Orbital_Ventures_SCA,_SICAV_RAIF_Schedule_of_Investments.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames[:5]:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    for i, row in enumerate(rows[:50]):
        vals = [str(v)[:25] if v is not None else '' for v in row[:10]]
        if any(v.strip() for v in vals):
            print(f"  {i+1}: {' | '.join(v for v in vals if v.strip())}")
