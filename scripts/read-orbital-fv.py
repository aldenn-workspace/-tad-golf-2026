import openpyxl

path = "/Users/mini/Downloads/2025_Orbital_Ventures_SCA,_SICAV_RAIF_Schedule_of_Investments.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

# Find the right sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    
    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v).lower() if v else '' for v in row]
        if any('fair value' in v or 'fmv' in v or 'cost' in v for v in vals):
            header_row = i
            print(f"\nSheet: {sheet_name} | Header row {i+1}:")
            print([str(v)[:20] if v else '' for v in row])
            break
    
    if header_row is not None:
        print(f"\nData rows ({sheet_name}):")
        for row in rows[header_row+1:header_row+25]:
            vals = [str(v)[:20] if v is not None else '' for v in row]
            if any(v.strip() for v in vals):
                print(' | '.join(v for v in vals if v.strip()))
