import openpyxl

wb = openpyxl.load_workbook("/Users/mini/Dropbox/MC with new Halter - Q4 2025 Combined SOI V2 -Draft.xlsx", read_only=True, data_only=True)

SOI_SHEETS = ['PV I SOI', 'PV II SOI', 'PV III SOI', 'PV E SOI', 'PV Whoop SOI', 'PVM Halter SOI', 'PVM Chef SOI']

for sheet_name in SOI_SHEETS:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print('='*60)
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Skip completely empty rows
        if any(v is not None for v in row):
            rows.append(row)
    # Print first 40 rows
    for i, row in enumerate(rows[:40]):
        # Filter out rows that are all None
        vals = [str(v) if v is not None else '' for v in row[:12]]
        if any(v.strip() for v in vals):
            print(f"  Row {i+1}: {' | '.join(v[:20] for v in vals if v.strip())}")
