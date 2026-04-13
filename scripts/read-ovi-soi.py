import openpyxl

path = "/Users/mini/Dropbox/OV Q4 2025 SOI Draft_PF_20260209_CPR 1.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames[:6]:
    ws = wb[sheet_name]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    print(f"\n{'='*70}")
    print(f"SHEET: {sheet_name}")
    for i, row in enumerate(rows[:60]):
        vals = [str(v)[:25] if v is not None else '' for v in row[:12]]
        if any(v.strip() for v in vals):
            print(f"  R{i+1}: " + ' | '.join(v for v in vals if v.strip()))
