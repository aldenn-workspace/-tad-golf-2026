import openpyxl

wb = openpyxl.load_workbook("/Users/mini/Dropbox/MC with new Halter - Q4 2025 Combined SOI V2 -Draft.xlsx", read_only=True, data_only=True)

SOI_SHEETS = ['PV I SOI', 'PV II SOI', 'PV III SOI', 'PV E SOI', 'PV Whoop SOI', 'PVM Halter SOI', 'PVM Chef SOI']

for sheet_name in SOI_SHEETS:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"SHEET: {sheet_name}")
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    for i, row in enumerate(rows[:80]):
        vals = [str(v)[:22] if v is not None else '' for v in row[:14]]
        if any(v.strip() for v in vals):
            print(f"  R{i+1}: " + ' | '.join(v for v in vals if v.strip()))
